#!/usr/bin/env python3
"""
station-fillins-sync.py
-----------------------
Watches station-fillins.xlsx and pushes its contents to a Google Sheet
whenever the file changes on disk. Runs on your Mac. Leave it running
in a terminal while Claude edits the xlsx; you'll see updates land live
in the Google Sheet in your browser.

Usage:
    python3 station-fillins-sync.py

Ctrl+C to stop.

Requirements:
    pip3 install gspread google-auth openpyxl
"""

from __future__ import annotations
import sys
import time
import traceback
from pathlib import Path

try:
    import gspread
    from google.oauth2.service_account import Credentials
    from openpyxl import load_workbook
except ImportError as e:
    print(f"Missing dependency: {e.name}")
    print("Install with:  pip3 install gspread google-auth openpyxl")
    sys.exit(1)


# ---------------- CONFIG ----------------
XLSX_PATH   = Path("/Users/raycheljohnson/Developer/CMS_April-2026/station-fillins-tabs.xlsx")
CREDS_PATH  = Path("/Users/raycheljohnson/Desktop/level6-node-factory/credentials.json")
SHEET_ID    = "1ul7NIOO-OhukZK0rx746_2n2wogODKfcXlFS8HwEU5Q"
POLL_SECONDS = 1.5
# ----------------------------------------

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def log(msg: str) -> None:
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)


def connect():
    creds = Credentials.from_service_account_file(str(CREDS_PATH), scopes=SCOPES)
    gc = gspread.authorize(creds)
    return gc.open_by_key(SHEET_ID)


def read_xlsx(path: Path) -> dict[str, list[list[str]]]:
    """Load workbook and return {sheet_name: 2D-string-grid}."""
    wb = load_workbook(str(path), data_only=True)
    out: dict[str, list[list[str]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        # fill merged-cell anchors so the top-left value is preserved
        merged_top_left = {}
        for mr in ws.merged_cells.ranges:
            tl = ws.cell(row=mr.min_row, column=mr.min_col).value
            if tl is not None:
                merged_top_left[(mr.min_row, mr.min_col)] = tl
        # find actual content extent
        max_r, max_c = 0, 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    if cell.row > max_r: max_r = cell.row
                    if cell.column > max_c: max_c = cell.column
        if max_r == 0:
            out[name] = []
            continue
        rows: list[list[str]] = []
        for r in range(1, max_r + 1):
            row: list[str] = []
            for c in range(1, max_c + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    v = ""
                row.append(str(v))
            rows.append(row)
        out[name] = rows
    return out


def sync(gs, data: dict[str, list[list[str]]]) -> None:
    existing = {ws.title: ws for ws in gs.worksheets()}
    first_tab = None

    for name, rows in data.items():
        if not rows:
            continue
        max_cols = max(len(r) for r in rows)
        # normalize row widths
        rows = [r + [""] * (max_cols - len(r)) for r in rows]

        if name not in existing:
            ws = gs.add_worksheet(
                title=name,
                rows=max(len(rows) + 10, 100),
                cols=max(max_cols + 5, 30),
            )
            log(f"  + created tab: {name}")
        else:
            ws = existing[name]
            # resize up if needed
            need_rows = max(len(rows) + 10, ws.row_count)
            need_cols = max(max_cols + 5, ws.col_count)
            if need_rows != ws.row_count or need_cols != ws.col_count:
                ws.resize(rows=need_rows, cols=need_cols)
            ws.clear()

        ws.update(values=rows, range_name="A1")
        log(f"  ✓ {name}: {len(rows)} rows × {max_cols} cols")
        if first_tab is None:
            first_tab = ws

    # clean up: delete the default "Sheet1" if it's empty and we've added real tabs
    if first_tab is not None:
        for ws in gs.worksheets():
            if ws.title == "Sheet1" and ws.title not in data:
                try:
                    gs.del_worksheet(ws)
                    log("  - removed empty default Sheet1")
                except Exception:
                    pass
                break


def main() -> None:
    if not CREDS_PATH.exists():
        sys.exit(f"Credentials not found: {CREDS_PATH}")
    if not XLSX_PATH.exists():
        sys.exit(f"xlsx not found: {XLSX_PATH}")

    log(f"Watching : {XLSX_PATH}")
    log(f"Target   : https://docs.google.com/spreadsheets/d/{SHEET_ID}/")

    try:
        gs = connect()
    except Exception as e:
        print("\n==== CONNECT FAILED ====")
        traceback.print_exc()
        print("\nCommon causes:")
        print("  1. The Sheet is not shared with the service account email.")
        print("     Open credentials.json and find `client_email` — share the Sheet")
        print("     with that address (Editor access).")
        print("  2. Google Sheets API or Drive API not enabled on the service")
        print("     account's Google Cloud project.")
        print("     Enable at: https://console.cloud.google.com/apis/library")
        sys.exit(1)
    log(f"Connected to Sheet: \"{gs.title}\"")

    last_mtime = 0.0
    while True:
        try:
            mtime = XLSX_PATH.stat().st_mtime
            if mtime != last_mtime:
                # small settle delay in case write is mid-flight
                time.sleep(0.3)
                try:
                    data = read_xlsx(XLSX_PATH)
                except Exception as e:
                    log(f"! parse error (will retry): {e}")
                    time.sleep(1)
                    continue
                log("change detected — syncing…")
                sync(gs, data)
                last_mtime = mtime
                log("done.\n")
            time.sleep(POLL_SECONDS)
        except KeyboardInterrupt:
            log("stopped.")
            break
        except Exception as e:
            log(f"! error: {e}")
            traceback.print_exc()
            time.sleep(5)
            try:
                gs = connect()
                log("reconnected.")
            except Exception:
                pass


if __name__ == "__main__":
    main()
